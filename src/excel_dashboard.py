"""Yes24 IT 모바일 베스트셀러 엑셀 대시보드 생성 스크립트.

CSV 데이터를 읽어 엑셀 파일에 차트, 요약 통계, 서식이 적용된
대시보드 형태로 저장합니다.
"""

import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ── 경로 설정 ──────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
CSV_PATH = DATA_DIR / "yes24_it_mobile_bestsellers.csv"
OUTPUT_PATH = DATA_DIR / "yes24_dashboard.xlsx"

# ── 스타일 상수 ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="맑은 고딕", bold=True, size=12, color="2E75B6")
METRIC_FONT = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
METRIC_LABEL_FONT = Font(name="맑은 고딕", size=10, color="666666")
BODY_FONT = Font(name="맑은 고딕", size=10)
LINK_FONT = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 차트 색상 팔레트
CHART_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B57A0", "636363", "EB7E30",
]


def _apply_header(ws, row: int, cols: int) -> None:
    """지정 행에 헤더 스타일을 적용한다."""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 50) -> None:
    """열 너비를 내용에 따라 자동 조절한다."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            # 한글은 약 2배 너비
            ko_chars = len(re.findall(r"[가-힣]", val))
            lengths.append(len(val) + ko_chars)
        best = max(lengths) + 2
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def _style_data_rows(ws, start_row: int, end_row: int, cols: int) -> None:
    """데이터 행에 기본 스타일과 교차 행 색상을 적용한다."""
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if c == 2 else CENTER  # 도서명만 왼쪽 정렬
        if (r - start_row) % 2 == 1:
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL


def load_data() -> pd.DataFrame:
    """CSV를 읽고 파생 컬럼을 추가한다."""
    df = pd.read_csv(CSV_PATH, encoding="utf-8-sig")

    # 가격 숫자화
    df["가격_숫자"] = (
        df["가격"].astype(str).str.replace(r"[^\d]", "", regex=True).str.strip()
    )
    df["가격_숫자"] = pd.to_numeric(df["가격_숫자"], errors="coerce")

    # 연도/월 추출
    df["출판연도"] = pd.to_numeric(
        df["출판일"].astype(str).str.extract(r"(\d{4})", expand=False),
        errors="coerce",
    )
    df["출판월"] = pd.to_numeric(
        df["출판일"].astype(str).str.extract(r"(\d{2})월", expand=False),
        errors="coerce",
    )
    return df


def create_dashboard(df: pd.DataFrame) -> None:
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    #  시트 1: 대시보드 요약
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "대시보드"
    ws.sheet_properties.tabColor = "1F4E79"

    # 제목
    ws.merge_cells("A1:H1")
    ws["A1"] = "Yes24 IT 모바일 베스트셀러 대시보드"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"총 {len(df):,}권  |  평균 가격 ₩{df['가격_숫자'].mean():,.0f}  |  출판사 {df['출판사'].nunique()}개  |  저자 {df['저자'].nunique()}명"
    ws["A2"].font = METRIC_LABEL_FONT
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 25

    # ── 요약 카드 (4개) ────────────────────────────────────────────────────
    metrics = [
        ("총 도서 수", f"{len(df):,}권", "A4"),
        ("평균 가격", f"₩{df['가격_숫자'].mean():,.0f}", "C4"),
        ("출판사 수", f"{df['출판사'].nunique()}개", "E4"),
        ("저자 수", f"{df['저자'].nunique()}명", "G4"),
    ]
    card_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for label, value, cell_ref in metrics:
        col_idx = ws[cell_ref].column
        ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 1)
        cell = ws[cell_ref]
        cell.value = value
        cell.font = METRIC_FONT
        cell.alignment = CENTER
        cell.fill = card_fill
        cell.border = THIN_BORDER
        # 레이블 행
        lbl_cell = ws.cell(row=5, column=col_idx)
        ws.merge_cells(start_row=5, start_column=col_idx, end_row=5, end_column=col_idx + 1)
        lbl_cell.value = label
        lbl_cell.font = METRIC_LABEL_FONT
        lbl_cell.alignment = CENTER
    ws.row_dimensions[4].height = 35

    # ── 출판사별 도서 수 (Top 10) ──────────────────────────────────────────
    ws["A7"] = "출판사별 도서 수 (Top 10)"
    ws["A7"].font = SUBTITLE_FONT
    ws.merge_cells("A7:E7")

    pub_counts = df["출판사"].value_counts().head(10).reset_index()
    pub_counts.columns = ["출판사", "도서 수"]
    start = 8
    ws.cell(row=start, column=1, value="출판사")
    ws.cell(row=start, column=2, value="도서 수")
    _apply_header(ws, start, 2)
    for i, (_, row) in enumerate(pub_counts.iterrows()):
        ws.cell(row=start + 1 + i, column=1, value=row["출판사"])
        ws.cell(row=start + 1 + i, column=2, value=int(row["도서 수"]))
    _style_data_rows(ws, start + 1, start + len(pub_counts), 2)

    # 막대 차트
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "출판사별 도서 수"
    chart1.y_axis.title = "출판사"
    chart1.x_axis.title = "도서 수"
    chart1.width = 22
    chart1.height = 13
    data1 = Reference(ws, min_col=2, min_row=start, max_row=start + len(pub_counts))
    cats1 = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(pub_counts))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.solidFill = CHART_COLORS[0]
    ws.add_chart(chart1, "D7")

    # ── 가격 분포 ──────────────────────────────────────────────────────────
    price_bins = [0, 15000, 20000, 25000, 30000, 50000]
    price_labels = ["~1.5만", "1.5~2만", "2~2.5만", "2.5~3만", "3만~"]
    df_price = df.dropna(subset=["가격_숫자"]).copy()
    df_price["가격대"] = pd.cut(df_price["가격_숫자"], bins=price_bins, labels=price_labels, right=False)
    price_dist = df_price["가격대"].value_counts().reindex(price_labels).fillna(0).astype(int)

    ws["A21"] = "가격대별 도서 수"
    ws["A21"].font = SUBTITLE_FONT
    ws.merge_cells("A21:E21")
    start2 = 22
    ws.cell(row=start2, column=1, value="가격대")
    ws.cell(row=start2, column=2, value="도서 수")
    _apply_header(ws, start2, 2)
    for i, label in enumerate(price_labels):
        ws.cell(row=start2 + 1 + i, column=1, value=label)
        ws.cell(row=start2 + 1 + i, column=2, value=int(price_dist[label]))
    _style_data_rows(ws, start2 + 1, start2 + len(price_labels), 2)

    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "가격대별 도서 수"
    chart2.width = 22
    chart2.height = 13
    data2 = Reference(ws, min_col=2, min_row=start2, max_row=start2 + len(price_labels))
    cats2 = Reference(ws, min_col=1, min_row=start2 + 1, max_row=start2 + len(price_labels))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.series[0].graphicalProperties.solidFill = CHART_COLORS[1]
    ws.add_chart(chart2, "D21")

    # ── 도서명 키워드 (Top 15) ─────────────────────────────────────────────
    all_words: list[str] = []
    for title in df["도서명"].dropna():
        tokens = re.findall(r"[가-힣]{2,}|[A-Za-z]{2,}", str(title))
        all_words.extend(tokens)
    stopwords = {
        "하는", "위한", "으로", "에서", "부터", "까지", "바로", "되는",
        "with", "the", "for", "and", "of", "to", "in", "Do", "it",
    }
    filtered = [w for w in all_words if w not in stopwords and len(w) > 1]
    kw_freq = Counter(filtered).most_common(15)

    ws["A30"] = "도서명 주요 키워드 (Top 15)"
    ws["A30"].font = SUBTITLE_FONT
    ws.merge_cells("A30:E30")
    start3 = 31
    ws.cell(row=start3, column=1, value="키워드")
    ws.cell(row=start3, column=2, value="빈도")
    _apply_header(ws, start3, 2)
    for i, (word, freq) in enumerate(kw_freq):
        ws.cell(row=start3 + 1 + i, column=1, value=word)
        ws.cell(row=start3 + 1 + i, column=2, value=freq)
    _style_data_rows(ws, start3 + 1, start3 + len(kw_freq), 2)

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.style = 10
    chart3.title = "도서명 주요 키워드"
    chart3.width = 22
    chart3.height = 13
    data3 = Reference(ws, min_col=2, min_row=start3, max_row=start3 + len(kw_freq))
    cats3 = Reference(ws, min_col=1, min_row=start3 + 1, max_row=start3 + len(kw_freq))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = CHART_COLORS[4]
    ws.add_chart(chart3, "D30")

    # ── 출판연도별 추이 ────────────────────────────────────────────────────
    year_counts = (
        df["출판연도"].dropna().astype(int).value_counts().sort_index().reset_index()
    )
    year_counts.columns = ["연도", "도서 수"]

    ws["A48"] = "출판연도별 도서 수"
    ws["A48"].font = SUBTITLE_FONT
    ws.merge_cells("A48:E48")
    start4 = 49
    ws.cell(row=start4, column=1, value="연도")
    ws.cell(row=start4, column=2, value="도서 수")
    _apply_header(ws, start4, 2)
    for i, (_, row) in enumerate(year_counts.iterrows()):
        ws.cell(row=start4 + 1 + i, column=1, value=int(row["연도"]))
        ws.cell(row=start4 + 1 + i, column=2, value=int(row["도서 수"]))
    _style_data_rows(ws, start4 + 1, start4 + len(year_counts), 2)

    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = "출판연도별 도서 수"
    chart4.width = 22
    chart4.height = 13
    data4 = Reference(ws, min_col=2, min_row=start4, max_row=start4 + len(year_counts))
    cats4 = Reference(ws, min_col=1, min_row=start4 + 1, max_row=start4 + len(year_counts))
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.series[0].graphicalProperties.solidFill = CHART_COLORS[5]
    ws.add_chart(chart4, "D48")

    # 열 너비
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14

    # ══════════════════════════════════════════════════════════════════════
    #  시트 2: 도서 목록
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("도서 목록")
    ws2.sheet_properties.tabColor = "2E75B6"

    headers = ["순위", "도서명", "저자", "출판사", "출판일", "가격", "링크"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _apply_header(ws2, 1, len(headers))
    ws2.row_dimensions[1].height = 25

    display_cols = ["순위", "도서명", "저자", "출판사", "출판일", "가격", "링크"]
    for r, (_, row) in enumerate(df[display_cols].iterrows(), 2):
        for c, col_name in enumerate(headers, 1):
            val = row[col_name]
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_name == "링크":
                cell.font = LINK_FONT
                cell.hyperlink = str(val)
                cell.alignment = CENTER
            elif col_name == "도서명":
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER
        if (r - 2) % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws2.cell(row=r, column=c).fill = ALT_ROW_FILL

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 15
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 48

    # 자동 필터
    ws2.auto_filter.ref = f"A1:G{len(df) + 1}"

    # ══════════════════════════════════════════════════════════════════════
    #  시트 3: 출판사별 상세
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("출판사별 분석")
    ws3.sheet_properties.tabColor = "ED7D31"

    pub_stats = (
        df.groupby("출판사")
        .agg(도서수=("도서명", "count"), 평균가격=("가격_숫자", "mean"), 최고가=("가격_숫자", "max"), 최저가=("가격_숫자", "min"))
        .sort_values("도서수", ascending=False)
        .reset_index()
    )
    pub_stats["평균가격"] = pub_stats["평균가격"].round(0).astype("Int64")
    pub_stats["최고가"] = pub_stats["최고가"].round(0).astype("Int64")
    pub_stats["최저가"] = pub_stats["최저가"].round(0).astype("Int64")

    pub_headers = ["출판사", "도서 수", "평균 가격", "최고 가격", "최저 가격"]
    for c, h in enumerate(pub_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    _apply_header(ws3, 1, len(pub_headers))
    ws3.row_dimensions[1].height = 25

    for r, (_, row) in enumerate(pub_stats.iterrows(), 2):
        ws3.cell(row=r, column=1, value=row["출판사"])
        ws3.cell(row=r, column=2, value=int(row["도서수"]))
        ws3.cell(row=r, column=3, value=row["평균가격"])
        ws3.cell(row=r, column=4, value=row["최고가"])
        ws3.cell(row=r, column=5, value=row["최저가"])
        for c in range(1, 6):
            cell = ws3.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if c in (3, 4, 5) and cell.value:
                cell.number_format = "#,##0"
        if (r - 2) % 2 == 1:
            for c in range(1, 6):
                ws3.cell(row=r, column=c).fill = ALT_ROW_FILL

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 14
    ws3.auto_filter.ref = f"A1:E{len(pub_stats) + 1}"

    # ══════════════════════════════════════════════════════════════════════
    #  저장
    # ══════════════════════════════════════════════════════════════════════
    wb.save(OUTPUT_PATH)
    print(f"엑셀 대시보드 생성 완료: {OUTPUT_PATH}")
    print(f"  - 대시보드 시트: 요약 통계 + 4개 차트")
    print(f"  - 도서 목록 시트: {len(df):,}권 (자동 필터 적용)")
    print(f"  - 출판사별 분석 시트: {len(pub_stats)}개 출판사")


def main() -> None:
    df = load_data()
    create_dashboard(df)


if __name__ == "__main__":
    main()
